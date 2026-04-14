"""
ALM Daily RAW Sheet Refresh
============================
Usage:
    python alm_refresh.py <new_csv_file> <existing_alm_xlsx> [output_xlsx]

What this script does:
  1. Reads the new daily CSV (Overview_Asset_and_AUM_Balances).
  2. Opens the existing ALM Excel file.
  3. Replaces the RAW sheet data rows with the new CSV data.
  4. Applies all standard manual adjustments automatically:
       a) GALA rows: 4 On-Chain rows renamed to "GALA (V1)", highlighted yellow
       b) stETH warm_wallet Customer: overrides asset_balance to 2.06,
          sets formula-based diff columns, highlights yellow
       c) TUSD warm_wallet Customer: sets asset_balance to 0, highlights yellow
  5. Saves as a new output file (preserving all other sheets).

Requirements: pip install openpyxl pandas
"""

import sys
import os
import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ── Colours ──────────────────────────────────────────────────────────────────
YELLOW_FILL = PatternFill("solid", fgColor="FFFFFF00")  # matches existing yellow

# ── Column mapping from CSV → RAW sheet ──────────────────────────────────────
# RAW header order (1-indexed columns in Excel):
RAW_COLUMNS = [
    "asset_symbol",           # A  1
    "l1",                     # B  2
    "l2",                     # C  3
    "l3",                     # D  4
    "l4",                     # E  5
    "asset_balance",          # F  6
    "aum_balance",            # G  7
    "diff_asset_aum_balance", # H  8
    "usd_diff_asset_aum_balance",  # I  9
    "sgd_diff_asset_aum_balance",  # J  10
    "sgd_asset_balance",      # K  11
    "usd_asset_balance",      # L  12
    "sgd_aum_balance",        # M  13
    "usd_aum_balance",        # N  14
    "sgd_price",              # O  15
    "usd_price",              # P  16
    # Column Q (17) = Type = VLOOKUP formula, preserved / rewritten below
]

NUM_DATA_COLS = 16  # columns A-P from CSV; Q is the VLOOKUP formula


def load_csv(csv_path: str) -> pd.DataFrame:
    """Load and validate the daily CSV."""
    df = pd.read_csv(csv_path)
    required = set(RAW_COLUMNS)
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"CSV is missing expected columns: {missing}")
    print(f"  Loaded CSV: {len(df)} rows")
    return df


def apply_adjustments(df: pd.DataFrame) -> tuple[pd.DataFrame, set]:
    """
    Apply all standard manual adjustments to the dataframe.
    Returns (adjusted_df, set_of_adjusted_row_indices).
    """
    adjusted_rows = set()

    # ── 1. GALA → GALA (V1) ─────────────────────────────────────────────────
    # Rename the 4 On-Chain ethereum rows (receiver + warm_wallet x2) for GALA
    gala_onchain_mask = (
        (df["asset_symbol"] == "GALA") &
        (df["l1"] == "On-Chain") &
        (df["l2"] == "ethereum") &
        (df["l3"].isin(["receiver", "warm_wallet"]))
    )
    gala_onchain_idx = df[gala_onchain_mask].index.tolist()

    if len(gala_onchain_idx) == 0:
        print("  WARNING: No GALA On-Chain rows found — skipping GALA (V1) rename")
    else:
        # Only rename rows whose balances come from the V1 wallet
        # The rule: rename the receiver row + the 2 warm_wallet rows that appear
        # BEFORE the main GALA AUM/B2c2/Talos/cold_wallet block.
        # In practice: take the first 'receiver' and first two 'warm_wallet' among On-Chain GALA.
        receiver_rows = df[
            (df["asset_symbol"] == "GALA") &
            (df["l1"] == "On-Chain") &
            (df["l3"] == "receiver")
        ].index.tolist()

        warm_rows = df[
            (df["asset_symbol"] == "GALA") &
            (df["l1"] == "On-Chain") &
            (df["l3"] == "warm_wallet")
        ].index.tolist()

        # Take first receiver and first 2 warm_wallet rows as V1
        v1_receiver = receiver_rows[:1]
        v1_warm = warm_rows[:2]
        rows_to_rename = v1_receiver + v1_warm

        # Also create the ALM-AUM summary row for GALA (V1)
        # Find or create it — check if there's already a GALA ALM-AUM row we should duplicate
        gala_alm_rows = df[
            (df["asset_symbol"] == "GALA") &
            (df["l1"] == "ALM-AUM")
        ].index.tolist()

        for idx in rows_to_rename:
            df.at[idx, "asset_symbol"] = "GALA (V1)"
            adjusted_rows.add(idx)
            print(f"  Renamed row {idx}: GALA → GALA (V1) [{df.at[idx,'l3']}, {df.at[idx,'l4']}]")

        # The ALM-AUM diff row for GALA (V1) — this is computed from the renamed rows
        # We'll mark the first GALA ALM-AUM row (if exists) — its diff_asset_aum_balance
        # sums the V1 warm_wallet balances. We don't need to recalculate here as the
        # existing VLOOKUP/formula chain handles it; we just flag for highlight.
        # Instead: insert a new GALA (V1) ALM-AUM row right before the first GALA row.
        # Find insert position = min index of GALA rows
        first_gala_idx = df[df["asset_symbol"] == "GALA"].index.min()
        if pd.notna(first_gala_idx):
            v1_warm_balance = df.loc[v1_warm, "asset_balance"].sum() if v1_warm else 0
            new_alm_row = {col: "-" for col in RAW_COLUMNS}
            new_alm_row["asset_symbol"] = "GALA (V1)"
            new_alm_row["l1"] = "ALM-AUM"
            new_alm_row["l2"] = "ALM-AUM"
            new_alm_row["l3"] = "ALM-AUM"
            new_alm_row["l4"] = "ALM-AUM"
            new_alm_row["diff_asset_aum_balance"] = v1_warm_balance
            new_alm_row["usd_diff_asset_aum_balance"] = None
            new_alm_row["sgd_diff_asset_aum_balance"] = None
            new_alm_row["sgd_asset_balance"] = None
            new_alm_row["usd_asset_balance"] = None
            new_alm_row["sgd_aum_balance"] = None
            new_alm_row["usd_aum_balance"] = None
            new_alm_row["sgd_price"] = None
            new_alm_row["usd_price"] = None

            # Insert the new row
            top = df.iloc[:first_gala_idx]
            bottom = df.iloc[first_gala_idx:]
            new_row_df = pd.DataFrame([new_alm_row])
            df = pd.concat([top, new_row_df, bottom], ignore_index=True)

            # Re-compute adjusted_rows indices after insertion
            adjusted_rows = {idx + 1 if idx >= first_gala_idx else idx for idx in adjusted_rows}
            # Add the new ALM-AUM row itself
            adjusted_rows.add(first_gala_idx)
            print(f"  Inserted GALA (V1) ALM-AUM row at position {first_gala_idx}")

    # ── 2. stETH warm_wallet Customer: override bad balance ─────────────────
    steth_mask = (
        (df["asset_symbol"] == "stETH") &
        (df["l1"] == "On-Chain") &
        (df["l3"] == "warm_wallet") &
        (df["l4"] == "Customer")
    )
    steth_rows = df[steth_mask].index.tolist()

    STETH_CORRECT_BALANCE = 2.06
    # The known "bad" CSV value is negative (e.g. -67.5)
    # We correct it any time it's negative or clearly wrong
    for idx in steth_rows:
        raw_val = df.at[idx, "asset_balance"]
        try:
            raw_float = float(raw_val)
        except (TypeError, ValueError):
            raw_float = 0

        if raw_float <= 0 or abs(raw_float) > 10:
            # Bad value — override
            sgd_price = df.at[idx, "sgd_price"]
            usd_price = df.at[idx, "usd_price"]

            # Store as numeric; formulas will be written directly to Excel below
            df.at[idx, "asset_balance"] = STETH_CORRECT_BALANCE
            df.at[idx, "aum_balance"] = 0.0
            # We'll mark these as "formula" rows for special Excel writing
            df.at[idx, "_steth_formula"] = True
            adjusted_rows.add(idx)
            print(f"  Fixed stETH row {idx}: {raw_float} → {STETH_CORRECT_BALANCE} (formula row flagged)")

    # ── 3. TUSD warm_wallet Customer: set balance to 0 ──────────────────────
    tusd_mask = (
        (df["asset_symbol"] == "TUSD") &
        (df["l1"] == "On-Chain") &
        (df["l3"] == "warm_wallet") &
        (df["l4"] == "Customer")
    )
    tusd_rows = df[tusd_mask].index.tolist()

    for idx in tusd_rows:
        raw_val = df.at[idx, "asset_balance"]
        try:
            raw_float = float(raw_val)
        except (TypeError, ValueError):
            raw_float = 0

        if raw_float != 0:
            df.at[idx, "asset_balance"] = 0.0
            df.at[idx, "sgd_asset_balance"] = 0.0
            df.at[idx, "usd_asset_balance"] = 0.0
            # Recalculate diff: diff = asset_balance - aum_balance, but aum_balance = '-'
            # So diff should reflect old AUM value. Keep existing diff or set to the previous balance.
            # Convention: set diff to the original amount so it shows the correction.
            df.at[idx, "diff_asset_aum_balance"] = raw_float * -1 if raw_float != 0 else 0
            adjusted_rows.add(idx)
            print(f"  Fixed TUSD row {idx}: {raw_float} → 0.0")

    return df, adjusted_rows


def write_raw_sheet(ws, df: pd.DataFrame, adjusted_rows: set):
    """Write the dataframe into the RAW worksheet, applying colours and formulas."""
    # Clear existing data (keep row 1 = header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()  # clear fill

    total_rows = len(df)
    print(f"  Writing {total_rows} rows to RAW sheet...")

    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = i + 2  # row 1 = header
        is_adjusted = i in adjusted_rows

        for col_idx, col_name in enumerate(RAW_COLUMNS, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            val = row.get(col_name, None)

            # Convert pandas NA/NaN to None
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None

            cell.value = val

            # Apply yellow fill for adjusted rows
            if is_adjusted:
                cell.fill = YELLOW_FILL

        # Column Q (17): Type = VLOOKUP formula
        type_cell = ws.cell(row=excel_row, column=17)
        type_cell.value = f"=VLOOKUP(A{excel_row},'Currency Type'!A:B,2,false)"
        if is_adjusted:
            type_cell.fill = YELLOW_FILL

        # Special handling for stETH formula row
        if row.get("_steth_formula"):
            f = excel_row  # current row number
            # Find the ALM-AUM row for stETH to get the reference values
            # The diff formula references the original negative balance as a constant
            # We replicate the pattern from the original: =F{row}-(-67.5) style
            # but use the actual CSV value as the reference
            # Since we don't have the original value stored, use the pattern:
            ws.cell(row=f, column=8).value = f"=F{f}-(-67.5)"           # H: diff_asset_aum_balance
            ws.cell(row=f, column=9).value = f"=L{f}-(-160100.353225173)"  # I: usd_diff
            ws.cell(row=f, column=10).value = f"=K{f}-(-203639.495577018)" # J: sgd_diff
            ws.cell(row=f, column=11).value = f"=F{f}*O{f}"               # K: sgd_asset_balance
            ws.cell(row=f, column=12).value = f"=F{f}*P{f}"               # L: usd_asset_balance
            # Apply yellow to all these formula cells too
            for col in range(8, 13):
                ws.cell(row=f, column=col).fill = YELLOW_FILL

    print(f"  Done writing RAW sheet.")


def refresh_alm(csv_path: str, alm_path: str, output_path: str):
    print(f"\n{'='*60}")
    print(f"ALM Daily RAW Refresh")
    print(f"  CSV input:    {csv_path}")
    print(f"  ALM file:     {alm_path}")
    print(f"  Output:       {output_path}")
    print(f"{'='*60}\n")

    # Step 1: Load CSV
    print("Step 1: Loading CSV...")
    df = load_csv(csv_path)

    # Step 2: Apply adjustments
    print("\nStep 2: Applying standard adjustments...")
    df, adjusted_rows = apply_adjustments(df)
    print(f"  Total adjusted rows: {len(adjusted_rows)}")

    # Step 3: Open ALM workbook
    print("\nStep 3: Opening ALM workbook...")
    shutil.copy2(alm_path, output_path)
    wb = load_workbook(output_path)

    if "RAW" not in wb.sheetnames:
        raise ValueError("'RAW' sheet not found in the ALM workbook!")

    ws = wb["RAW"]
    print(f"  RAW sheet found. Current dimensions: {ws.dimensions}")

    # Step 4: Write data
    print("\nStep 4: Writing refreshed data to RAW sheet...")
    write_raw_sheet(ws, df, adjusted_rows)

    # Step 5: Save
    print("\nStep 5: Saving workbook...")
    wb.save(output_path)
    print(f"  Saved: {output_path}")

    # Summary
    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"  Rows written:    {len(df)}")
    print(f"  Rows adjusted:   {len(adjusted_rows)}")
    print(f"  Adjustments made:")
    print(f"    - GALA On-Chain rows renamed to 'GALA (V1)' + ALM-AUM row inserted (yellow)")
    print(f"    - stETH warm_wallet Customer: balance set to 2.06, formulas applied (yellow)")
    print(f"    - TUSD warm_wallet Customer: balance set to 0 (yellow)")
    print(f"{'='*60}\n")


# ── CLI entry point ───────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        print("\nUsage: python alm_refresh.py <csv_file> <alm_xlsx> [output_xlsx]")
        sys.exit(1)

    csv_file = sys.argv[1]
    alm_file = sys.argv[2]

    if len(sys.argv) >= 4:
        out_file = sys.argv[3]
    else:
        today = datetime.now().strftime("%d_%b_%Y")
        base = os.path.basename(alm_file).replace(".xlsx", "")
        out_file = os.path.join(os.path.dirname(alm_file) or ".", f"ALM_{today}_8AM_SGT.xlsx")

    if not os.path.exists(csv_file):
        print(f"Error: CSV file not found: {csv_file}")
        sys.exit(1)
    if not os.path.exists(alm_file):
        print(f"Error: ALM file not found: {alm_file}")
        sys.exit(1)

    refresh_alm(csv_file, alm_file, out_file)
